from io import BytesIO
from openpyxl import Workbook


def create_excel(survey_id, period, submission_list):
    """Extract comments from submissions and write them to an excel file"""
    print("Generating Excel file")
    workbook = Workbook()
    row = 2
    surveys_with_comments_count = len(submission_list)
    ws = workbook.active

    for submission in submission_list:
        comment = submission["comment"]

        boxes_selected = submission["boxes_selected"]

        if not comment:
            continue
        row += 1
        surveys_with_comments_count += 1
        ws.cell(row, 1, submission['ru_ref'])
        ws.cell(row, 2, period)

        if survey_id == '134':
            additional_list = submission["additional"]
            for addition in additional_list:
                if addition['qcode'] == '300w':
                    ws.cell(row, 5, addition['comment'])
                if addition['qcode'] == '300f':
                    ws.cell(row, 6, addition['comment'])
                if addition['qcode'] == '300m':
                    ws.cell(row, 7, addition['comment'])
                if addition['qcode'] == '300w4':
                    ws.cell(row, 8, addition['comment'])
                if addition['qcode'] == '300w5':
                    ws.cell(row, 9, addition['comment'])

        ws.cell(row, 3, boxes_selected)
        ws.cell(row, 4, comment)

    ws.cell(1, 1, f"Survey ID: {survey_id}")
    ws.cell(1, 2, f"Comments found: {surveys_with_comments_count}")
    if survey_id == '134':
        ws.cell(1, 5, "Weekly comment")
        ws.cell(1, 6, "Fortnightly comment")
        ws.cell(1, 7, "Calendar Monthly comment")
        ws.cell(1, 8, "4 Weekly Pay comment")
        ws.cell(1, 9, "5 Weekly Pay comment")
    print(f"{surveys_with_comments_count} out of {len(submission_list)} submissions had comments")

    filename = "test.xlsx"
    # workbook.save(filename)
    workbook.close()

    print(f"Excel file {filename} generated")

    virtual_workbook = BytesIO()
    workbook.save(virtual_workbook)

    return filename, virtual_workbook.getvalue()


# create_excel("009", "2019", [submission2, submission2])

