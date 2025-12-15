from openpyxl import Workbook
from io import BytesIO

from openpyxl.utils.exceptions import IllegalCharacterError
from openpyxl.worksheet.worksheet import Worksheet

from app import get_logger
from app.submission import Submission


logger = get_logger()


class ExcelWriter:

    def create_excel(self, survey_id: str, submission_list: list[Submission]):
        """
        Generates an Excel file from the list of comments passed in that relate to the given survey_id and period.
        Returns the file and the count of comments within the file.
        """
        logger.info("Generating Excel file")
        workbook = Workbook()
        row_index = 3
        surveys_with_comments_count = 0
        ws = workbook.active

        for submission in submission_list:
            if submission.has_comment():
                self.add_row(ws, row_index, survey_id, submission)
                row_index += 1
                surveys_with_comments_count += 1

        # header
        ws.cell(1, 1, f"Survey ID: {survey_id}")
        ws.cell(1, 2, f"Comments found: {surveys_with_comments_count}")
        if survey_id == '134':
            ws.cell(1, 5, "Weekly comment")
            ws.cell(1, 6, "Fortnightly comment")
            ws.cell(1, 7, "Calendar Monthly comment")
            ws.cell(1, 8, "4 Weekly Pay comment")
            ws.cell(1, 9, "5 Weekly Pay comment")
        logger.info(f"{surveys_with_comments_count} out of {len(submission_list)} submissions had comments")

        workbook.close()

        virtual_workbook = BytesIO()
        workbook.save(virtual_workbook)

        return virtual_workbook.getvalue(), surveys_with_comments_count

    def add_row(self, ws: Worksheet, row_index: int, survey_id: str, submission: Submission):
        ws.cell(row_index, 1, submission.ru_ref)
        ws.cell(row_index, 2, submission.period)

        additional_mappings = {'300w': 5, '300f': 6, '300m': 7, '300w4': 8, '300w5': 9}
        if survey_id == '134':
            for qcode, additional_comment in submission.additional.items():
                if qcode in additional_mappings:
                    try:
                        ws.cell(row_index, additional_mappings[qcode], additional_comment)
                    except IllegalCharacterError:
                        logger.info("Comment with illegal character found")
                        ws.cell(row_index, additional_mappings[qcode], 'This comment contained illegal characters that are not printable')

        ws.cell(row_index, 3, submission.boxes_selected)
        try:
            ws.cell(row_index, 4, submission.comment)
        except IllegalCharacterError:
            logger.info("Comment with illegal character found")
            ws.cell(row_index, 4, 'This comment contained illegal characters that are not printable')
